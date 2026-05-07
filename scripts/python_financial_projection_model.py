#!/usr/bin/env python3
"""Create financial_projections_final from financial_projections.

The script copies the source workbook to a separate final workbook path with
``shutil.copy2`` and then applies user-provided assumption changes only to the
copied workbook. This keeps ``data/financial_projections.xlsx`` as the
source-of-truth template while letting you generate a changed
``financial_projections_final.xlsx`` without opening Excel manually.

Examples:
    python scripts/python_financial_projection_model.py --list-assumptions
    python scripts/python_financial_projection_model.py --set "Fuel expense per operating car (monthly)=12000"
    python scripts/python_financial_projection_model.py --assumptions-file scenario.json
    python scripts/python_financial_projection_model.py --interactive
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Iterable

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_INPUT = Path("data/financial_projections.xlsx")
DEFAULT_OUTPUT_NAME = "financial_projections_final.xlsx"
ASSUMPTIONS_SHEET_NAME = "Assumptions"

COST_OF_SALES_LABEL = "Cost of sales"
MONTHLY_GROSS_REVENUE_LABEL = "Monthly Gross revenue per car"
MONTHLY_GROSS_PROFIT_LABEL = "Monthly Gross profit per car"
SALARY_EXPENSE_LABEL = "Salary expense per operating car (monthly)"
TOTAL_OPERATING_EXPENSES_LABEL = "Total operating expenses per operating car (monthly)"
OPERATING_PROFIT_LABEL = "Operating Profit"


@dataclass(frozen=True)
class AssumptionRow:
    """A discovered assumption row in the Assumptions worksheet."""

    label: str
    value_cell: str
    value: object
    units: object
    notes: object


def copy_financial_projection_workbook(input_path: Path, output_path: Path) -> None:
    """Copy the source workbook to the output path before any modifications."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_path, output_path)


def normalize_assumption_key(value: str) -> str:
    """Normalize an assumption label for forgiving CLI lookups."""
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def cell_display_value(value: object) -> str | None:
    """Convert an openpyxl cell value to display text for CLI output."""
    if value is None:
        return None
    return str(value)


def require_assumptions_sheet(workbook: object) -> Worksheet:
    """Return the Assumptions worksheet from an open workbook."""
    if ASSUMPTIONS_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain a {ASSUMPTIONS_SHEET_NAME!r} sheet.")
    return workbook[ASSUMPTIONS_SHEET_NAME]


def list_assumptions_from_worksheet(worksheet: Worksheet) -> list[AssumptionRow]:
    """Read assumptions from an open Assumptions worksheet."""
    assumptions: list[AssumptionRow] = []

    for row_index in range(1, worksheet.max_row + 1):
        label = worksheet.cell(row=row_index, column=1).value
        if not label or label in {"Assumption", "Namibia Rent-to-Own Fleet Model - Key Assumptions (Editable)"}:
            continue

        assumptions.append(
            AssumptionRow(
                label=str(label),
                value_cell=f"B{row_index}",
                value=worksheet.cell(row=row_index, column=2).value,
                units=worksheet.cell(row=row_index, column=3).value,
                notes=worksheet.cell(row=row_index, column=4).value,
            )
        )

    return assumptions


def list_assumptions(workbook_path: Path) -> list[AssumptionRow]:
    """Read assumptions from the workbook's Assumptions sheet with openpyxl."""
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False, keep_links=True, rich_text=True)
    worksheet = require_assumptions_sheet(workbook)
    assumptions = list_assumptions_from_worksheet(worksheet)
    workbook.close()
    return assumptions


def parse_assumption_value(raw_value: object) -> object:
    """Convert user input to a value that openpyxl can safely write to a cell."""
    if isinstance(raw_value, bool | int | float):
        return raw_value

    value = str(raw_value).strip()
    if not value:
        raise ValueError("Assumption values cannot be blank.")
    if value.startswith("="):
        return value

    normalized = value.replace(",", "")
    try:
        numeric_value = float(normalized)
    except ValueError:
        return value

    if numeric_value.is_integer() and re.fullmatch(r"[-+]?\d+", normalized):
        return int(normalized)
    return numeric_value


def set_cell_value(cell: Cell, raw_value: object) -> None:
    """Set a worksheet cell to a numeric, boolean, string, or formula value."""
    cell.value = parse_assumption_value(raw_value)


def find_target_assumption(key: str, assumption_rows: list[AssumptionRow]) -> AssumptionRow:
    """Find an assumption by exact label, normalized label, or value cell."""
    stripped_key = key.strip()
    by_exact = {row.label: row for row in assumption_rows}
    by_cell = {row.value_cell.upper(): row for row in assumption_rows}
    by_normalized = {normalize_assumption_key(row.label): row for row in assumption_rows}

    if re.fullmatch(r"B\d+", stripped_key, flags=re.IGNORECASE):
        cell_key = stripped_key.upper()
        if cell_key in by_cell:
            return by_cell[cell_key]
        available_cells = ", ".join(row.value_cell for row in assumption_rows)
        raise KeyError(f"Unknown assumption value cell {key!r}. Available value cells: {available_cells}")

    if stripped_key in by_exact:
        return by_exact[stripped_key]

    normalized_key = normalize_assumption_key(stripped_key)
    if normalized_key in by_normalized:
        return by_normalized[normalized_key]

    available = ", ".join(row.label for row in assumption_rows)
    raise KeyError(f"Unknown assumption {key!r}. Available assumptions: {available}")


def write_direct_assumption_updates(
    worksheet: Worksheet,
    assumption_rows: list[AssumptionRow],
    updates: dict[str, object],
) -> dict[str, str]:
    """Apply explicit user-provided assumption updates to the worksheet."""
    applied: dict[str, str] = {}

    for key, value in updates.items():
        target_row = find_target_assumption(key, assumption_rows)
        set_cell_value(worksheet[target_row.value_cell], value)
        applied[target_row.label] = str(value)

    return applied


def read_numeric_assumption_values(worksheet: Worksheet, assumption_rows: list[AssumptionRow]) -> dict[str, float]:
    """Read numeric assumption values after edits."""
    values: dict[str, float] = {}
    for assumption in assumption_rows:
        raw = worksheet[assumption.value_cell].value
        if raw is None or isinstance(raw, bool):
            continue
        if isinstance(raw, str) and raw.startswith("="):
            continue
        try:
            values[assumption.label] = float(str(raw).replace(",", ""))
        except ValueError:
            continue
    return values


def update_calculated_assumption_cells(
    worksheet: Worksheet,
    assumption_rows: list[AssumptionRow],
    explicit_update_labels: set[str] | None = None,
) -> dict[str, float]:
    """Refresh known calculated assumption rows unless users set them directly."""
    explicit_update_labels = explicit_update_labels or set()
    numeric_values = read_numeric_assumption_values(worksheet, assumption_rows)
    cost_of_sales_inputs = [
        "Fuel expense per operating car (monthly)",
        "Airtime expense per operating car (monthly)",
        "Carwash expense per operating car (monthly)",
        "Maintenance expense per active car (monthly)",
        "Driver subsistence",
        "Incidental repair reserve",
        "Tracking device expense",
    ]

    calculated_updates: dict[str, float] = {}
    calculated_values: dict[str, float] = {}

    if all(label in numeric_values for label in cost_of_sales_inputs):
        if COST_OF_SALES_LABEL in explicit_update_labels and COST_OF_SALES_LABEL in numeric_values:
            calculated_values[COST_OF_SALES_LABEL] = numeric_values[COST_OF_SALES_LABEL]
        else:
            calculated_values[COST_OF_SALES_LABEL] = sum(
                numeric_values[label] for label in cost_of_sales_inputs
            )
            calculated_updates[COST_OF_SALES_LABEL] = calculated_values[COST_OF_SALES_LABEL]

    if MONTHLY_GROSS_REVENUE_LABEL in numeric_values and COST_OF_SALES_LABEL in calculated_values:
        if (
            MONTHLY_GROSS_PROFIT_LABEL in explicit_update_labels
            and MONTHLY_GROSS_PROFIT_LABEL in numeric_values
        ):
            calculated_values[MONTHLY_GROSS_PROFIT_LABEL] = numeric_values[MONTHLY_GROSS_PROFIT_LABEL]
        else:
            calculated_values[MONTHLY_GROSS_PROFIT_LABEL] = (
                numeric_values[MONTHLY_GROSS_REVENUE_LABEL] - calculated_values[COST_OF_SALES_LABEL]
            )
            calculated_updates[MONTHLY_GROSS_PROFIT_LABEL] = calculated_values[
                MONTHLY_GROSS_PROFIT_LABEL
            ]

    if SALARY_EXPENSE_LABEL in numeric_values and COST_OF_SALES_LABEL in calculated_values:
        if (
            TOTAL_OPERATING_EXPENSES_LABEL in explicit_update_labels
            and TOTAL_OPERATING_EXPENSES_LABEL in numeric_values
        ):
            calculated_values[TOTAL_OPERATING_EXPENSES_LABEL] = numeric_values[
                TOTAL_OPERATING_EXPENSES_LABEL
            ]
        else:
            calculated_values[TOTAL_OPERATING_EXPENSES_LABEL] = (
                calculated_values[COST_OF_SALES_LABEL] + numeric_values[SALARY_EXPENSE_LABEL]
            )
            calculated_updates[TOTAL_OPERATING_EXPENSES_LABEL] = calculated_values[
                TOTAL_OPERATING_EXPENSES_LABEL
            ]

    if MONTHLY_GROSS_PROFIT_LABEL in calculated_values and SALARY_EXPENSE_LABEL in numeric_values:
        if OPERATING_PROFIT_LABEL in explicit_update_labels and OPERATING_PROFIT_LABEL in numeric_values:
            calculated_values[OPERATING_PROFIT_LABEL] = numeric_values[OPERATING_PROFIT_LABEL]
        else:
            calculated_values[OPERATING_PROFIT_LABEL] = (
                calculated_values[MONTHLY_GROSS_PROFIT_LABEL] - numeric_values[SALARY_EXPENSE_LABEL]
            )
            calculated_updates[OPERATING_PROFIT_LABEL] = calculated_values[OPERATING_PROFIT_LABEL]

    if calculated_updates:
        write_direct_assumption_updates(worksheet, assumption_rows, calculated_updates)
    return calculated_updates


def force_full_workbook_recalculation(workbook: object) -> None:
    """Ask Excel-compatible applications to recalculate formulas when opened."""
    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def validate_xlsx(workbook_path: Path) -> None:
    """Validate that the saved output is a readable XLSX archive and workbook."""
    from openpyxl import load_workbook

    with zipfile.ZipFile(workbook_path) as workbook_archive:
        bad_file = workbook_archive.testzip()
        if bad_file is not None:
            raise ValueError(f"Generated workbook is corrupted; first bad archive member: {bad_file}")

    validation_workbook = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=True)
    validation_workbook.close()


def apply_assumption_updates(workbook_path: Path, updates: dict[str, object]) -> dict[str, str]:
    """Apply assumption updates in-place to an xlsx workbook copy with openpyxl."""
    if not updates:
        return {}

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False, keep_links=True, rich_text=True)
    worksheet = require_assumptions_sheet(workbook)
    assumption_rows = list_assumptions_from_worksheet(worksheet)
    applied = write_direct_assumption_updates(worksheet, assumption_rows, updates)
    calculated = update_calculated_assumption_cells(
        worksheet, assumption_rows, explicit_update_labels=set(applied)
    )

    force_full_workbook_recalculation(workbook)
    workbook.save(workbook_path)
    workbook.close()
    validate_xlsx(workbook_path)

    return applied | {label: str(value) for label, value in calculated.items() if label not in applied}


def parse_set_arguments(set_arguments: Iterable[str]) -> dict[str, str]:
    """Parse repeated --set KEY=VALUE arguments."""
    updates: dict[str, str] = {}
    for item in set_arguments:
        if "=" not in item:
            raise ValueError(f"Invalid --set value {item!r}. Use --set 'Assumption name=123'.")
        key, value = item.split("=", 1)
        key = key.strip()
        if not key:
            raise ValueError(f"Invalid --set value {item!r}. Assumption name cannot be blank.")
        updates[key] = value.strip()
    return updates


def load_assumption_file(path: Path) -> dict[str, object]:
    """Load assumption overrides from a JSON file."""
    with path.open("r", encoding="utf-8") as file_handle:
        data = json.load(file_handle)
    if not isinstance(data, dict):
        raise ValueError("Assumptions file must contain a JSON object of assumption names to values.")
    return data


def prompt_for_assumptions(workbook_path: Path) -> dict[str, str]:
    """Collect assumption overrides from standard input."""
    print("Available assumptions:")
    for assumption in list_assumptions(workbook_path):
        units = f" ({assumption.units})" if assumption.units else ""
        print(f"- {assumption.label}{units}: current value = {cell_display_value(assumption.value)}")

    print("\nEnter changes as Assumption name=value. Press Enter on a blank line when done.")
    updates: dict[str, str] = {}
    while True:
        try:
            line = input("> ").strip()
        except EOFError:
            break
        if not line:
            break
        if "=" not in line:
            print("Please use Assumption name=value format.")
            continue
        key, value = line.split("=", 1)
        updates[key.strip()] = value.strip()
    return updates


def print_assumption_list(workbook_path: Path) -> None:
    """Print editable assumptions for CLI users."""
    for assumption in list_assumptions(workbook_path):
        units = f" [{assumption.units}]" if assumption.units else ""
        value = cell_display_value(assumption.value)
        print(f"{assumption.label} ({assumption.value_cell}) = {value}{units}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate financial_projections_final by first copying the source workbook "
            "with shutil.copy2, then modifying only the copied workbook with openpyxl. "
            "Use the optional input layer to override assumptions."
        )
    )
    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT),
        help="Path to the source financial_projections workbook.",
    )
    parser.add_argument(
        "--output",
        help="Path to the output financial_projections_final workbook.",
    )
    parser.add_argument(
        "--set",
        dest="set_values",
        action="append",
        default=[],
        metavar="ASSUMPTION=VALUE",
        help=(
            "Override one assumption in the generated final workbook. Repeat for "
            "multiple changes. Use the exact assumption name from --list-assumptions "
            "or a value cell such as B8."
        ),
    )
    parser.add_argument(
        "--assumptions-file",
        help="JSON file containing assumption-name-to-value overrides.",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Prompt for assumption changes before generating the final workbook.",
    )
    parser.add_argument(
        "--list-assumptions",
        action="store_true",
        help="List editable assumptions from the input workbook and exit.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    if args.list_assumptions:
        print_assumption_list(input_path)
        return

    updates: dict[str, object] = {}
    if args.assumptions_file:
        updates.update(load_assumption_file(Path(args.assumptions_file)))
    updates.update(parse_set_arguments(args.set_values))
    if args.interactive:
        updates.update(prompt_for_assumptions(input_path))

    output_path = Path(args.output) if args.output else input_path.with_name(DEFAULT_OUTPUT_NAME)
    copy_financial_projection_workbook(input_path=input_path, output_path=output_path)
    applied_updates = apply_assumption_updates(output_path, updates)

    print(f"Financial projections final workbook generated: {output_path}")
    if applied_updates:
        print("Applied assumption updates:")
        for label, value in applied_updates.items():
            print(f"- {label}: {value}")


if __name__ == "__main__":
    main()
